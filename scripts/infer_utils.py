import numpy as np
import cv2
import tensorflow as tf

import fitz
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment

import math
import os
import sys

sys.path.insert(0, os.path.join("..", "keras_unets"))

# sys.path.insert(0, "../models/CRAFT-pytorch")

# from craft import CRAFT
# import torch

from keras_unet_collection.models import att_unet_2d

TEXT_DETECTION_MEAN = (0.485, 0.456, 0.406)
TEXT_DETECTION_VARIANCE = (0.229, 0.224, 0.225)
MAX_VALUE = 255

TABLE_DETECTION_CONFIG = {
    "input_shape" : (512, 512),
    "band_size" : 2
}

TEXT_DETECTION_CONFIG = {
    "input_shape" : (600, 800),
    "band_size" : 3
}

# initialize models

# text detector model

# torch_craft = CRAFT()
# torch_craft.load_state_dict(torch.load("../models/CRAFT-pytorch/craft_mlt_25k.pth", map_location=torch.device('cpu')), strict=False)

text_detector_interpreter = tf.lite.Interpreter("../models/lite-model_craft-text-detector_float16_1.tflite")
text_detector_interpreter.allocate_tensors()
text_detector_input_details = text_detector_interpreter.get_input_details()
text_detector_output_details = text_detector_interpreter.get_output_details()

# table detector model
down_scales = [32, 64, 128, 256]
table_detector_model = att_unet_2d((TABLE_DETECTION_CONFIG["input_shape"][0], TABLE_DETECTION_CONFIG["input_shape"][1], 2), down_scales, n_labels=1,
                            stack_num_down=2, stack_num_up=2,
                            activation='ReLU', atten_activation='ReLU', attention='add', output_activation="Sigmoid", 
                            batch_norm=True, pool=False, unpool='bilinear', name='attunet'
                        )

checkpoint = tf.train.Checkpoint(step=tf.Variable(1), optimizer=tf.keras.optimizers.Adam(), net=table_detector_model)
# print(f"loading checkpoint {'training_checkpoints/' + '2022.07.30-22/' + 'ckpt-238'}")
status = checkpoint.restore("training_checkpoints/2022.08.29-07/ckpt-157")
status.expect_partial()

# end initialize models

"""  
Copyright (c) 2019-present NAVER Corp.
MIT License
"""

""" auxilary functions """
# unwarp corodinates
def warpCoord(Minv, pt):
    out = np.matmul(Minv, (pt[0], pt[1], 1))
    return np.array([out[0]/out[2], out[1]/out[2]])
""" end of auxilary functions """


def getDetBoxes_core(textmap, linkmap, text_threshold, link_threshold, low_text):
    # prepare data
    linkmap = linkmap.copy()
    textmap = textmap.copy()
    img_h, img_w = textmap.shape

    """ labeling method """
    ret, text_score = cv2.threshold(textmap, low_text, 1, 0)
    ret, link_score = cv2.threshold(linkmap, link_threshold, 1, 0)

    text_score_comb = np.clip(text_score + link_score, 0, 1)
    nLabels, labels, stats, centroids = cv2.connectedComponentsWithStats(text_score_comb.astype(np.uint8), connectivity=4)

    det = []
    mapper = []
    for k in range(1,nLabels):
        # size filtering
        size = stats[k, cv2.CC_STAT_AREA]
        if size < 10: continue

        # thresholding
        if np.max(textmap[labels==k]) < text_threshold: continue

        # make segmentation map
        segmap = np.zeros(textmap.shape, dtype=np.uint8)
        segmap[labels==k] = 255
        segmap[np.logical_and(link_score==1, text_score==0)] = 0   # remove link area
        x, y = stats[k, cv2.CC_STAT_LEFT], stats[k, cv2.CC_STAT_TOP]
        w, h = stats[k, cv2.CC_STAT_WIDTH], stats[k, cv2.CC_STAT_HEIGHT]
        niter = int(math.sqrt(size * min(w, h) / (w * h)) * 2)
        sx, ex, sy, ey = x - niter, x + w + niter + 1, y - niter, y + h + niter + 1
        # boundary check
        if sx < 0 : sx = 0
        if sy < 0 : sy = 0
        if ex >= img_w: ex = img_w
        if ey >= img_h: ey = img_h
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT,(1 + niter, 1 + niter))
        segmap[sy:ey, sx:ex] = cv2.dilate(segmap[sy:ey, sx:ex], kernel)

        # make box
        np_contours = np.roll(np.array(np.where(segmap!=0)),1,axis=0).transpose().reshape(-1,2)
        rectangle = cv2.minAreaRect(np_contours)
        box = cv2.boxPoints(rectangle)

        # align diamond-shape
        w, h = np.linalg.norm(box[0] - box[1]), np.linalg.norm(box[1] - box[2])
        box_ratio = max(w, h) / (min(w, h) + 1e-5)
        if abs(1 - box_ratio) <= 0.1:
            l, r = min(np_contours[:,0]), max(np_contours[:,0])
            t, b = min(np_contours[:,1]), max(np_contours[:,1])
            box = np.array([[l, t], [r, t], [r, b], [l, b]], dtype=np.float32)

        # make clock-wise order
        startidx = box.sum(axis=1).argmin()
        box = np.roll(box, 4-startidx, 0)
        box = np.array(box)

        det.append(box)
        mapper.append(k)

    return det, labels, mapper

def getPoly_core(boxes, labels, mapper, linkmap):
    # configs
    num_cp = 5
    max_len_ratio = 0.7
    expand_ratio = 1.45
    max_r = 2.0
    step_r = 0.2

    polys = []  
    for k, box in enumerate(boxes):
        # size filter for small instance
        w, h = int(np.linalg.norm(box[0] - box[1]) + 1), int(np.linalg.norm(box[1] - box[2]) + 1)
        if w < 10 or h < 10:
            polys.append(None); continue

        # warp image
        tar = np.float32([[0,0],[w,0],[w,h],[0,h]])
        M = cv2.getPerspectiveTransform(box, tar)
        word_label = cv2.warpPerspective(labels, M, (w, h), flags=cv2.INTER_NEAREST)
        try:
            Minv = np.linalg.inv(M)
        except:
            polys.append(None); continue

        # binarization for selected label
        cur_label = mapper[k]
        word_label[word_label != cur_label] = 0
        word_label[word_label > 0] = 1

        """ Polygon generation """
        # find top/bottom contours
        cp = []
        max_len = -1
        for i in range(w):
            region = np.where(word_label[:,i] != 0)[0]
            if len(region) < 2 : continue
            cp.append((i, region[0], region[-1]))
            length = region[-1] - region[0] + 1
            if length > max_len: max_len = length

        # pass if max_len is similar to h
        if h * max_len_ratio < max_len:
            polys.append(None); continue

        # get pivot points with fixed length
        tot_seg = num_cp * 2 + 1
        seg_w = w / tot_seg     # segment width
        pp = [None] * num_cp    # init pivot points
        cp_section = [[0, 0]] * tot_seg
        seg_height = [0] * num_cp
        seg_num = 0
        num_sec = 0
        prev_h = -1
        for i in range(0,len(cp)):
            (x, sy, ey) = cp[i]
            if (seg_num + 1) * seg_w <= x and seg_num <= tot_seg:
                # average previous segment
                if num_sec == 0: break
                cp_section[seg_num] = [cp_section[seg_num][0] / num_sec, cp_section[seg_num][1] / num_sec]
                num_sec = 0

                # reset variables
                seg_num += 1
                prev_h = -1

            # accumulate center points
            cy = (sy + ey) * 0.5
            cur_h = ey - sy + 1
            cp_section[seg_num] = [cp_section[seg_num][0] + x, cp_section[seg_num][1] + cy]
            num_sec += 1

            if seg_num % 2 == 0: continue # No polygon area

            if prev_h < cur_h:
                pp[int((seg_num - 1)/2)] = (x, cy)
                seg_height[int((seg_num - 1)/2)] = cur_h
                prev_h = cur_h

        # processing last segment
        if num_sec != 0:
            cp_section[-1] = [cp_section[-1][0] / num_sec, cp_section[-1][1] / num_sec]

        # pass if num of pivots is not sufficient or segment widh is smaller than character height 
        if None in pp or seg_w < np.max(seg_height) * 0.25:
            polys.append(None); continue

        # calc median maximum of pivot points
        half_char_h = np.median(seg_height) * expand_ratio / 2

        # calc gradiant and apply to make horizontal pivots
        new_pp = []
        for i, (x, cy) in enumerate(pp):
            dx = cp_section[i * 2 + 2][0] - cp_section[i * 2][0]
            dy = cp_section[i * 2 + 2][1] - cp_section[i * 2][1]
            if dx == 0:     # gradient if zero
                new_pp.append([x, cy - half_char_h, x, cy + half_char_h])
                continue
            rad = - math.atan2(dy, dx)
            c, s = half_char_h * math.cos(rad), half_char_h * math.sin(rad)
            new_pp.append([x - s, cy - c, x + s, cy + c])

        # get edge points to cover character heatmaps
        isSppFound, isEppFound = False, False
        grad_s = (pp[1][1] - pp[0][1]) / (pp[1][0] - pp[0][0]) + (pp[2][1] - pp[1][1]) / (pp[2][0] - pp[1][0])
        grad_e = (pp[-2][1] - pp[-1][1]) / (pp[-2][0] - pp[-1][0]) + (pp[-3][1] - pp[-2][1]) / (pp[-3][0] - pp[-2][0])
        for r in np.arange(0.5, max_r, step_r):
            dx = 2 * half_char_h * r
            if not isSppFound:
                line_img = np.zeros(word_label.shape, dtype=np.uint8)
                dy = grad_s * dx
                p = np.array(new_pp[0]) - np.array([dx, dy, dx, dy])
                cv2.line(line_img, (int(p[0]), int(p[1])), (int(p[2]), int(p[3])), 1, thickness=1)
                if np.sum(np.logical_and(word_label, line_img)) == 0 or r + 2 * step_r >= max_r:
                    spp = p
                    isSppFound = True
            if not isEppFound:
                line_img = np.zeros(word_label.shape, dtype=np.uint8)
                dy = grad_e * dx
                p = np.array(new_pp[-1]) + np.array([dx, dy, dx, dy])
                cv2.line(line_img, (int(p[0]), int(p[1])), (int(p[2]), int(p[3])), 1, thickness=1)
                if np.sum(np.logical_and(word_label, line_img)) == 0 or r + 2 * step_r >= max_r:
                    epp = p
                    isEppFound = True
            if isSppFound and isEppFound:
                break

        # pass if boundary of polygon is not found
        if not (isSppFound and isEppFound):
            polys.append(None); continue

        # make final polygon
        poly = []
        poly.append(warpCoord(Minv, (spp[0], spp[1])))
        for p in new_pp:
            poly.append(warpCoord(Minv, (p[0], p[1])))
        poly.append(warpCoord(Minv, (epp[0], epp[1])))
        poly.append(warpCoord(Minv, (epp[2], epp[3])))
        for p in reversed(new_pp):
            poly.append(warpCoord(Minv, (p[2], p[3])))
        poly.append(warpCoord(Minv, (spp[2], spp[3])))

        # add to final result
        polys.append(np.array(poly))

    return polys

def getDetBoxes(textmap, linkmap, text_threshold, link_threshold, low_text, poly=False):
    boxes, labels, mapper = getDetBoxes_core(textmap, linkmap, text_threshold, link_threshold, low_text)

    if poly:
        polys = getPoly_core(boxes, labels, mapper, linkmap)
    else:
        polys = [None] * len(boxes)

    return boxes, polys

def adjustResultCoordinates(polys, ratio_w, ratio_h, ratio_net = 2):
    if len(polys) > 0:
        polys = np.array(polys)
        for k in range(len(polys)):
            if polys[k] is not None:
                polys[k] *= (ratio_w * ratio_net, ratio_h * ratio_net)
    return polys

def cvt2HeatmapImg(img):
    img = (np.clip(img, 0, 1) * 255).astype(np.uint8)
    img = cv2.applyColorMap(img, cv2.COLORMAP_JET)
    return img

# end cited functions


# helper functions

def normalize_table_detector_input(img, resize_shape):
    img = cv2.resize(img, resize_shape, cv2.INTER_AREA)

    edges = cv2.bitwise_not(cv2.Canny(img, 1, 10))
    img = np.moveaxis(np.array([img, edges]), 0, -1)

    return tf.convert_to_tensor(img / MAX_VALUE, dtype=tf.float32)


def normalize_table_detector_inputs(images, resize_shape):
    batch_X = []
    for img in images:
        img = cv2.resize(img, resize_shape, cv2.INTER_AREA)

        edges = cv2.bitwise_not(cv2.Canny(img, 1, 10))
        img = np.moveaxis(np.array([img, edges]), 0, -1)

        batch_X.append(img / MAX_VALUE)

    return tf.convert_to_tensor(batch_X, dtype=tf.float32)


def postprocess_table_detector_output(raw, min_pixel_size, min_area, max_seg_dist=0):

    rgb_mask = np.array(raw * 255, dtype=np.uint8)

    thresh = thresh = cv2.threshold(rgb_mask, 250, 255, cv2.THRESH_BINARY)[1]
    contours, hierarchy = cv2.findContours(image=thresh, mode=cv2.RETR_TREE, method=cv2.CHAIN_APPROX_SIMPLE)
    
    pred = np.zeros_like(raw, dtype=np.uint8)
    seg_coords = []

    for ind, c in enumerate(contours):
        if len(c) > min_pixel_size:
            min_x, max_x = np.squeeze(c)[:, 0].min(), np.squeeze(c)[:, 0].max()
            min_y, max_y = np.squeeze(c)[:, 1].min(), np.squeeze(c)[:, 1].max()
            if (max_x - min_x) * (max_y - min_y) > min_area:
                to_be_removed = []
                put_min_x, put_min_y, put_max_x, put_max_y = min_x, min_y, max_x, max_y
                for i, coords in enumerate(seg_coords):
                    x1, y1 = max(0, put_min_x - (max_seg_dist // 2)), max(0, put_min_y - (max_seg_dist // 2))
                    x2, y2 = max(0, put_max_x + (max_seg_dist // 2)), max(0, put_max_y + (max_seg_dist // 2))
                    if do_intersect((x1, y1, x2, y2), coords):
                        
                        put_min_x = min(x1, coords[0])
                        put_min_y = min(y1, coords[1])
                        put_max_x = max(x2, coords[2])
                        put_max_y = max(y2, coords[3])
                        to_be_removed.append(coords)
                        # print("inter", coords, (put_min_y, put_max_y, put_min_x, put_max_x))

                for rem in to_be_removed:
                    seg_coords.remove(rem)
                seg_coords.append((put_min_x, put_min_y, put_max_x, put_max_y))
                pred[put_min_y:put_max_y, put_min_x:put_max_x] = 255

    return pred


def normalize_text_detector_input(img, resize_shape):
    img = cv2.resize(img, resize_shape, cv2.INTER_AREA)

    img = np.array([img, img, img], np.uint8)
    # print(fill_img.shape)
    img = np.array([
                (img[0] - TEXT_DETECTION_MEAN[0] * 255) / (TEXT_DETECTION_VARIANCE[0] * 255),
                (img[1] - TEXT_DETECTION_MEAN[1] * 255) / (TEXT_DETECTION_VARIANCE[1] * 255),
                (img[2] - TEXT_DETECTION_MEAN[2] * 255) / (TEXT_DETECTION_VARIANCE[2] * 255)
            ])

    return tf.convert_to_tensor(img, tf.float32)


def normalize_text_detector_inputs(images, resize_shape):
    batch_X = []
    for img in images:
        resized_image = cv2.resize(img, resize_shape, cv2.INTER_AREA)

        fill_img = np.moveaxis(resized_image, -1, 0)
        # print(fill_img.shape)
        fill_img = np.array([
                    (fill_img[0] - TEXT_DETECTION_MEAN[0] * 255) / (TEXT_DETECTION_VARIANCE[0] * 255),
                    (fill_img[1] - TEXT_DETECTION_MEAN[1] * 255) / (TEXT_DETECTION_VARIANCE[1] * 255),
                    (fill_img[2] - TEXT_DETECTION_MEAN[2] * 255) / (TEXT_DETECTION_VARIANCE[2] * 255)
                ])
        batch_X.append(fill_img)

    return tf.convert_to_tensor(batch_X, tf.float32)


def do_intersect(r1, r2, margins=(0, 0)):
    r1 = r1[0] - margins[0], r1[1] - margins[1], r1[2] + margins[0], r1[3] + margins[1]
    r2 = r2[0] - margins[0], r2[1] - margins[1], r2[2] + margins[0], r2[3] + margins[1]
    return not (
        r2[0] > r1[2] or 
        r2[2] < r1[0] or 
        r2[1] > r1[3] or
        r2[3] < r1[1]
    )


def get_row_bboxes_count(bboxes):
    cur = bboxes[0]
    count = 1
    min_y = cur[3]

    for i in range(1, len(bboxes)):
        cur = bboxes[i]
        if cur[1] > min_y:
            break
        count += 1
        min_y = min(min_y, cur[3])

    return count


def get_col_bboxes_count(bboxes):
    cur = bboxes[0]
    count = 1
    max_x = cur[2]

    for i in range(1, len(bboxes)):
        cur = bboxes[i]
        if cur[0] > max_x:
            break
        count += 1
        max_x = min(max_x, cur[2])

    return count


def rescale_output(coords, current_shape, target_shape):
    # print(current_size, target_size)
    scale_x = target_shape[1]/current_shape[1]
    scale_y = target_shape[0]/current_shape[0]

    rescaled_coords = []
    for i in range(len(coords)):
        rescaled_coords.append([
            coords[i][0] * scale_x,
            coords[i][1] * scale_y,
            coords[i][2] * scale_x,
            coords[i][3] * scale_y
        ])

    return rescaled_coords


def read_pdf_windowed(fitz_doc_page, bbox, embrace=(5, 10)):
    return fitz_doc_page.get_textbox(
        fitz.Rect(bbox[0] - embrace[0], bbox[1] - embrace[1], bbox[2] + embrace[0], bbox[3] + embrace[1])
    )


def draw_table_struct(image, bboxes, name, sample_name):

    res_img = cv2.cvtColor(image, cv2.COLOR_GRAY2RGB)
    color_list = [(0, 0, 255), (0, 255, 0), (255, 0, 0), (0, 255, 255), (255, 255, 0), (255, 0, 255)]

    bboxes.sort(key=lambda x : x[0])

    total_count = 0
    color_index = 0
    column_midpoints = []

    while total_count < len(bboxes):
        count = get_col_bboxes_count(bboxes[total_count:])

        column_midpoint = round(np.array([(b[0], b[2]) for b in bboxes[total_count:total_count+count]]).mean(axis=1).mean())
        column_midpoints.append(column_midpoint)
        # print(column_midpoint)

        # for box in bboxes[total_count:total_count+count]:
        #     cv2.rectangle(res_img, (box[0] - 5, box[1] - 5), (box[2] + 5, box[3] + 5), color_list[color_index], thickness=1)

        cv2.line(res_img, (column_midpoint, 0), (column_midpoint, res_img.shape[0]), color=color_list[color_index], thickness=2)

        total_count += count

        color_index += 1
        if color_index == len(color_list) - 1:
            color_index = 0

    bboxes.sort(key=lambda x:x[1])

    total_count = 0
    while total_count < len(bboxes):
        count = get_row_bboxes_count(bboxes[total_count:])

        for box in bboxes[total_count:total_count+count]:
            cv2.rectangle(res_img, (round(box[0]), round(box[1])), (round(box[2]), round(box[3])), color_list[color_index], thickness=1)

        total_count += count

        color_index += 1
        if color_index == len(color_list) - 1:
            color_index = 0

    print("Svaed table struct ", f"../res/preds/{name}/{sample_name}.png")
    cv2.imwrite(f"../res/preds/{name}/{sample_name}.png", res_img)


def to_excel_file(pdf_bboxes, fits_doc_page, doc_name, sample_name):
    # create excel sheet
    workbook = Workbook()
    sheet = workbook.active

    # for i in range(len(bboxes)):
    # bboxes.sort(key=lambda x:get_contour_row_precedence(x, img.shape[1]))
    pdf_bboxes.sort(key=lambda x : x[0])

    total_count = 0
    column_midpoints = []

    while total_count < len(pdf_bboxes):
        count = get_col_bboxes_count(pdf_bboxes[total_count:])

        column_midpoint = np.array([(b[0], b[2]) for b in pdf_bboxes[total_count:total_count+count]]).mean(axis=1).mean()
        column_midpoints.append(column_midpoint)
        total_count += count

    pdf_bboxes.sort(key=lambda x:x[1])

    total_count = 0
    row_index = 1
    while total_count < len(pdf_bboxes):
        count = get_row_bboxes_count(pdf_bboxes[total_count:])

        sorted_row_bboxes = sorted(pdf_bboxes[total_count:total_count+count], key=lambda x : x[0])

        for box in sorted_row_bboxes:
            # intersected_column_indices = []
            minimal_dist = np.inf
            minimal_dist_column_ind = None
            # print("box", box)
            for col_ind, col_mid in enumerate(column_midpoints):
                # print("col mid", col_mid)
                distance_from_midpoint = abs((box[0] + box[2]) / 2 - col_mid)
                if minimal_dist > distance_from_midpoint:
                    minimal_dist = distance_from_midpoint
                    minimal_dist_column_ind = col_ind+1
                # if box[0] <= col_mid and box[2] >= col_mid:
                #     intersected_column_indices.append(col_ind+1)

            # if len(intersected_column_indices) >= 2:
            #     print("Merging cells.")
            #     sheet.merge_cells(
            #         start_row=row_index, 
            #         start_column=min(intersected_column_indices), 
            #         end_row=row_index, 
            #         end_column=max(intersected_column_indices)
            #     )

            if minimal_dist_column_ind:
                cell = sheet.cell(row_index, minimal_dist_column_ind)
                cell.value = read_pdf_windowed(fits_doc_page, box)
                cell.alignment = Alignment(horizontal="center")
            # else:
            #     cell = sheet.cell(row_index, min(intersected_column_indices))

            # if cell:
                # print(row_index, intersected_column_indices, minimal_dist_column_ind)
                # print("Writing...", read_pdf_windowed(fits_doc_page, box))
                
        for box in sorted_row_bboxes:
            intersected_column_indices = []

            for col_ind, col_mid in enumerate(column_midpoints):
                if box[0] <= col_mid and box[2] >= col_mid:
                    intersected_column_indices.append(col_ind+1)

            if len(intersected_column_indices) >= 2:
                sheet.merge_cells(
                    start_row=row_index, 
                    start_column=min(intersected_column_indices), 
                    end_row=row_index, 
                    end_column=max(intersected_column_indices)
                )

        row_index += 1
        total_count += count
    
    workbook.save(filename=f"../res/excel/{doc_name}/{sample_name}.xlsx")

# end helper functions


def detect_text_bboxes(input_image):

    # print(tf.expand_dims(input_image, 0).shape)
    import time
    st = time.time()

    text_detector_interpreter.set_tensor(text_detector_input_details[0]['index'], tf.expand_dims(input_image, 0))
    text_detector_interpreter.invoke()

    output_data = text_detector_interpreter.get_tensor(text_detector_output_details[0]['index'])

    print("Text pred took", time.time() - st, "secs")

    # with torch.no_grad():
    #     output_data, feature = torch_craft(torch.unsqueeze(torch.Tensor(np.array(input_image)), 0))

    #     print(output_data.shape, feature.shape)

    score_text = np.array(output_data[0,:,:,0])
    score_link = np.array(output_data[0,:,:,1])
    
    boxes, polys = getDetBoxes(score_text, score_link, 0.7, 0.4, 0.4, False)

    boxes = adjustResultCoordinates(boxes, 1, 1)
    
    valid_boxes = []

    for i, pred_box in enumerate(boxes):

        poly = np.array(pred_box).astype(np.int32)

        expand_size = 3
        b1 = [poly[0][0], poly[0][1], poly[2][0], poly[2][1]]
        b1[0] = b1[0] * input_image.shape[1] / input_image.shape[2]
        b1[2] = b1[2] * input_image.shape[1] / input_image.shape[2]
        b1[1] = b1[1] * input_image.shape[2] / input_image.shape[1]
        b1[3] = b1[3] * input_image.shape[2] / input_image.shape[1]

        # b1[0] -= expand_size
        # b1[2] += expand_size
        
        to_be_removed = []
        x1, y1, x2, y2 = b1[0], b1[1], b1[2], b1[3]
        for j, b2 in enumerate(valid_boxes):
            
            if do_intersect((x1, y1, x2, y2), b2, margins=(5, 0)):
                x1 = min(b2[0], x1)
                y1 = min(b2[1], y1)
                x2 = max(b2[2], x2)
                y2 = max(b2[3], y2)
                # valid_boxes.remove(pred_box)
                to_be_removed.append(b2)

        for rem in to_be_removed:
            valid_boxes.remove(rem)

        # print("fk rect", (x1, y1, x2, y2))
        valid_boxes.append((x1, y1, x2, y2))

    return valid_boxes


def detect_table_bboxes(input_image, pdf_name=None, sample_name=None):

    raw_out = tf.squeeze(table_detector_model(tf.expand_dims(input_image, 0), training=False))

    # to be deleted
    # print(np.min(input_image), np.max(input_image))
    inp_img = cv2.cvtColor(np.array(input_image[:, :, 0] * 255, dtype=np.uint8), cv2.COLOR_GRAY2RGB)

    process_output = postprocess_table_detector_output(raw_out, 2, 3000, max_seg_dist=40)
    # process_output = postprocess_table_detector_output(process_output, 2, 1000, max_seg_dist=20)

    contours, hierarchy = cv2.findContours(process_output, mode=cv2.RETR_TREE, method=cv2.CHAIN_APPROX_SIMPLE)

    valid_bboxes = []

    for ind, c in enumerate(contours):
        min_x, max_x = np.squeeze(c)[:, 0].min(), np.squeeze(c)[:, 0].max()
        min_y, max_y = np.squeeze(c)[:, 1].min(), np.squeeze(c)[:, 1].max()
        cv2.rectangle(inp_img, (min_x, min_y), (max_x, max_y), (200, 120, 0), thickness=2)
        valid_bboxes.append((min_x, min_y, max_x, max_y))

    if pdf_name:
        # print(inp_img.min(), inp_img.max())
        cv2.imwrite(
            f"../res/masks/{pdf_name}/{sample_name}.png",
            cv2.hconcat([
                inp_img,
                cv2.cvtColor(np.array(raw_out * 255, dtype=np.uint8), cv2.COLOR_GRAY2RGB),
                cv2.cvtColor(np.array(process_output, dtype=np.uint8), cv2.COLOR_GRAY2RGB)
            ])
        )

    return valid_bboxes